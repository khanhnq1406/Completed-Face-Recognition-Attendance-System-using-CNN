
from openpyxl import Workbook, load_workbook
from datetime import datetime
# wb=load_workbook('E:/test/namedb.xlsx')
# ws=wb.active
# col=ws['A']
# ID='A' + '1'
# ws[ID].value="Test"
# wb.save('E:/test/namedb.xlsx')
# name=[]
# for cell in col:
#     name.append(cell.value)
#     print(cell)
# print(name)

wb1 = load_workbook('../database/att_db.xlsx')
ws1 = wb1.active
# name=ws1['A']
# timein=ws1['B']
now = datetime.now()
dtString = now.strftime('%H:%M:%S')

for count in range(1, 100):
    name = 'A' + str(count)
    if ws1[name].value is None:
        ws1[name] = 'Khanh'
        wb1.save('../database/att_db.xlsx')
        break
    else:
        pass