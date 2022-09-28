from tkinter import *
from tkinter.ttk import Progressbar
from tkinter import messagebox
from subprocess import call
import cv2
from openpyxl import Workbook,load_workbook
import pathlib
import os
import pandas as pd
import openpyxl
# from FaceReg.face_dataset import count
#import FaceReg.face_training


global face_id,count
def newattendace():
    def getName():
        wb = load_workbook('../database/namedb.xlsx')
        ws = wb.active
        idplus=int(face_id) + 1
        idtrans=str(idplus)
        ID = 'A' +  idtrans
        ws[ID].value = name
        wb.save('../database/namedb.xlsx')
    def getinfo():
        global face_id,name
        face_id = id_entry.get()
        name=name_entry.get()
        getName()
        # call(["python", "FaceReg/face_dataset.py"])
        takeimg()
        #progressing()

    '''---Window---'''
    from main import mainfunc
    newattwindow = Canvas(mainfunc().window, width=mainfunc().width_value, height=mainfunc().height_value)
    newattwindow.place(x=0,y=0)
    # newattwindow.title("New Face Attendance")
    # newattwindow.geometry("960x540")
    # newattwindow.config(background="#323436")

    '''---Lable---'''
    id_lable = Label(newattwindow,
                      text='Enter your ID:',
                      font=('Montserrat', 25),
                      fg="#9aa9a9",
                      bg='#323436')
    name_lable = Label(newattwindow,
                      text='Enter your name:',
                      font=('Montserrat', 25),
                      fg="#9aa9a9",
                      bg='#323436')
    id_lable.place(x=10,y=10)
    name_lable.place(x=10,y=150)
    '''---Entry--'''
    id_entry=Entry(newattwindow,
                       font=('Montserrat', 25),
                       bg='#2b2c2e',
                       fg='#52acc7')
    name_entry = Entry(newattwindow,
                         font=('Montserrat', 25),
                         bg='#2b2c2e',
                         fg='#52acc7')
    id_entry.place(x=350,y=10)
    name_entry.place(x=350,y=150)

    '''---Button---'''

    submitname = Button(newattwindow,
                        text='Submit',
                        command=getinfo,
                        font=('Montserrat', 15),
                        height='2',
                        width='15',
                        fg='#9aa9a9',
                        bg='#252527',
                        )
    submitname.place(x=610, y=250)
    def progressing():
        progress_lable = Label(newattwindow,
                               text='Initializing face capture. Look the camera and wait',
                               font=('Montserrat', 12),
                               fg="#9aa9a9",
                               bg='#323436')
        progress_lable.place(x=80, y=350)
        progress = Progressbar(newattwindow, orient=HORIZONTAL, length=800)
        progress.place(x=80, y=380)
        progress['value'] += count
        newattwindow.update_idletasks()

    def takeimg():

        cam = cv2.VideoCapture(0)

        # Dùng thuật toán xác định khuôn mặt haarcascade_frontalface
        face_detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

        print("\n [INFO] Initializing face capture. Look the camera and wait ...")

        # Initialize individual sampling face count
        global count
        count = 0

        while (True):

            ret, img = cam.read()
            img = cv2.flip(img, 1)  # flip video image vertically
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_detector.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y), (x + w + 50, y + h + 50), (255, 0, 0), 2)
                count += 1  # sau mỗi lần lấy ảnh thành công, biến count sẽ tăng 1

                # Lưu hình ảnh đã chụp vào folder "dataset"
                gray = gray[y:y + h, x:x + w]

                # Lưu ảnh với định dạng "User.id.count.jpg
                cv2.imwrite("training/dataset/User." + str(face_id) + '.' + str(count) + ".jpg", gray)

                cv2.imshow('Taking Face Samples', img)
            progressing()
            k = cv2.waitKey(100) & 0xff
            if k == 27:
                break
            elif count >= 100:
                break
        cam.release()
        cv2.destroyAllWindows()
        successtakeface()
        training_lable = Label(newattwindow,
                               text='Training your face',
                               font=('Montserrat', 12),
                               fg="#9aa9a9",
                               bg='#323436')
        training_lable.place(x=80, y=430)
        training = Progressbar(newattwindow, orient=HORIZONTAL, length=800)
        training.place(x=80, y=460)
        for i in range(0,1000):
            training['value'] += 0.01
            newattwindow.update_idletasks()
        call(["python", "training/face_training.py"])
        for i in range(0,1000):
            training['value'] += 0.1
            newattwindow.update_idletasks()
        successtraining()
        newattwindow.destroy()
    def successtakeface():
        messagebox.showinfo(title='Done',message='Completed your face capture!!! Click OK to start the training process')
    def successtraining():
        messagebox.showinfo(title='Done',
                            message='Completed face training!!!\n Thank you '+str(name))


