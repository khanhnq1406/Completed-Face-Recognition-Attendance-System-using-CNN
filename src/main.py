import tkinter
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Progressbar
from tkinter import messagebox
from subprocess import call
# from newattendace import *
# from attendance import *
import cv2
from openpyxl import Workbook,load_workbook
import os
from PIL import Image,ImageTk
from datetime import datetime
import training.Model
import numpy as np
import sqlite3
import tkintertable
from tkintertable import TableCanvas, TableModel
from chuyenanhgray import *
global face_id, count


def mainfunc():
    '''---------Cửa sổ chính---------'''
    window = Tk()   #Tạo cửa sổ window
    width_value = 1920
    height_value = 1080
    window.geometry("%dx%d+0+0" % (width_value, height_value))  #Thiết lập kích thước cửa sổ
    window.title('Group 7 - Face Attendance System')    #Tên cửa sổ
    icon = PhotoImage(file='../pic/logospkt.png')  #icon của cửa sổ
    window.iconphoto(True,icon)

    def inmain():
        '''---------Tạo canvas cho màn hình HOME---------'''
        c1 = Canvas(window, width=width_value, height=height_value) #Thiết lập canvas có kích thước 1920x1080
        c1.place(x=0, y=0)
        menu_side = Label(c1,bg='#1C2A3A', padx=150, pady=1080) #Tạo lable menu
        menu_side.place(x=0, y=0)
        logo=ImageTk.PhotoImage(Image.open('../pic/logo77.jpg'))
        lg=Label(c1,image=logo)
        lg.image=logo
        lg.place(x=27,y=27)
        Label(c1,text='Face Attendance',font=('Montserrat', 15),bg='#1C2A3A',fg='#ffffff').place(x=110,y=20)
        Label(c1, text='System', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=50)
        Label(c1, text='Group 7', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=80)
        bg_photo=Image.open('../pic/FirstPage_bd0_f0.png') #Mở hình ảnh banner ở trang màn hình chính
        bg_photo=bg_photo.resize((1620,912),Image.BICUBIC)  #Thay đổi kích thước banner thành 1620x912 với bộ lọc BICUBIC
        bg_photo = ImageTk.PhotoImage(bg_photo)
        bg = tkinter.Label(image=bg_photo,bd=0) #Tạo lable cho banner với vị trí ơ305;60]
        bg.image=bg_photo
        bg.place(x=305,y=60)

        '''---------Hàm đăng ký thành viên mới---------'''
        def newattendace():
            def getName():
                # wb1 = load_workbook('E:/test/namedb.xlsx')
                # ws1 = wb1.active
                # idplus = int(face_id) + 1
                # idtrans = str(idplus)
                # ID = 'A' + idtrans
                # ws1[ID].value = name
                # wb1.save('E:/test/namedb.xlsx')
                '''Database'''
                conn = sqlite3.connect('../database/info.db')
                c = conn.cursor()
                c.execute('INSERT INTO info (id,name,username,password) values (?,?,?,?)',
                          (int(face_id), name, username, password))
                conn.commit()
                conn.close()
            def getinfo():
                global face_id, name,username,password
                face_id = id_entry.get()
                name = name_entry.get()
                username=username_entry.get()
                password=password_entry.get()
                getName()
                # call(["python", "training/face_dataset.py"])
                takeimg()
                # progressing()

            '''---Tạo canvas cho giao diện New Attendance---'''

            newattwindow = Canvas(window,width=width_value,height=height_value)
            newattwindow.place(x=0,y=0)

            '''Menu'''

            menu_side = Label(newattwindow, bg='#1C2A3A', padx=150, pady=1080)
            menu_side.place(x=0, y=0)
            logo = ImageTk.PhotoImage(Image.open('logo77.jpg'))
            lg = Label(newattwindow, image=logo)
            lg.image = logo
            lg.place(x=27, y=27)
            Label(newattwindow, text='Face Attendance', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=20)
            Label(newattwindow, text='System', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=50)
            Label(newattwindow, text='Group 7', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=80)

            '''---Lable---'''
            id_lable = Label(newattwindow,
                             text='ID:',
                             font=('Montserrat Light', 20),
                             fg="#000000",
                             bg='#f0f0f0')
            name_lable = Label(newattwindow,
                               text='Name:',
                               font=('Montserrat Light', 20),
                               fg="#000000",
                               bg='#f0f0f0')
            username_lable = Label(newattwindow,
                             text='Username:',
                             font=('Montserrat Light', 20),
                             fg="#000000",
                             bg='#f0f0f0')
            password_lable = Label(newattwindow,
                               text='Password:',
                               font=('Montserrat Light', 20),
                               fg="#000000",
                               bg='#f0f0f0')

            cam_photo = Image.open('../pic/Attendance_Picture.png')
            cam_photo = ImageTk.PhotoImage(cam_photo)
            lable_image = Label(newattwindow, image=cam_photo)
            lable_image.image=cam_photo
            lable_image.place(x=1200, y=50)

            id_lable.place(x=400, y=50)
            name_lable.place(x=400, y=170)
            username_lable.place(x=400, y=270)
            password_lable.place(x=400, y=390)
            # username_lable.place(x=1200, y=50)
            # password_lable.place(x=1200, y=170)
            '''---Entry--'''
            id_entry = Entry(newattwindow,
                             font=('Montserrat', 20),
                             bg='#ffffff',
                             fg='#000000')
            name_entry = Entry(newattwindow,
                               font=('Montserrat', 20),
                               bg='#ffffff',
                               fg='#000000')
            username_entry = Entry(newattwindow,
                             font=('Montserrat', 20),
                             bg='#ffffff',
                             fg='#000000')
            password_entry = Entry(newattwindow,
                               font=('Montserrat', 20),
                               bg='#ffffff',
                               fg='#000000',
                               show="*")
            id_entry.place(x=405, y=95)
            name_entry.place(x=405, y=215)
            username_entry.place(x=405, y=315)
            password_entry.place(x=405, y=435)
            # username_entry.place(x=1205, y=95)
            # password_entry.place(x=1205, y=215)
            '''---Button---'''

            submitname = Button(newattwindow,
                                text='Submit',
                                command=getinfo,
                                font=('Montserrat', 15),
                                height='1',
                                width='15',
                                fg='#ffffff',
                                bg='#2B3647',
                                )
            submitname.place(x=585, y=500)
            home = Button(
                window,
                text='Home',
                command=inmain,
                font=('Montserrat', 15),
                height='2',
                width='22',
                fg='#ffffff',
                bg='#1C2A3A',
                bd=0
            )
            newatt = Button(window,
                            text='Register New Employee',
                            command=newattendace,
                            font=('Montserrat', 15),
                            height='2',
                            width='23',
                            fg='#ffffff',
                            bg='#1FB29E',
                            bd=0
                            )
            att = Button(window,
                         text='Mark Attendance',
                         font=('Montserrat', 15),
                         command=attendance,
                         height='2',
                         width='22',
                         fg='#ffffff',
                         bg='#1C2A3A',
                         bd=0
                         )
            result_att = Button(window,
                            text='Result Attendance',
                            font=('Montserrat', 15),
                            command=result_attendance,
                            height='2',
                            width='22',
                            fg='#ffffff',
                            bg='#1C2A3A',
                            bd=0
                            )
            home.place(x=0, y=150)
            newatt.place(x=0, y=250)
            att.place(x=0, y=350)
            result_att.place(x=0, y=450)

            def progressing():
                progress_lable = Label(newattwindow,
                                       text='Initializing face capture. Look the camera and wait',
                                       font=('Montserrat', 12),
                                       fg="#000000",
                                       bg='#f0f0f0')
                progress_lable.place(x=400, y=610)
                progress = Progressbar(newattwindow, orient=HORIZONTAL, length=1450)
                progress.place(x=400, y=640)
                progress['value'] += count
                newattwindow.update_idletasks()

            '''---------Hàm lấy dataset---------'''
            def takeimg():

                cam = cv2.VideoCapture(0)

                # Dùng thuật toán xác định khuôn mặt haarcascade_frontalface
                face_detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

                print("\n [INFO] Initializing face capture. Look the camera and wait ...")

                # Khai báo biến đếm số lượng khuôn mặt sẽ được chụp
                global count
                count = 0

                while (True):
                    ret, img = cam.read()
                    img = cv2.flip(img, 1)  # flip video image vertically
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    faces = face_detector.detectMultiScale(gray, 1.3, 5)
                    for (x, y, w, h) in faces:
                        cv2.rectangle(img, (x, y), (x + w + 50, y + h + 50), (255, 0, 0), 2)
                        count += 1  # sau mỗi lần lấy ảnh thành công, biến count sẽ tăng 1

                        # Lưu hình ảnh đã chụp vào folder "dataset"
                        gray = gray[y:y + h, x:x + w]

                        # Lưu ảnh với định dạng "User.id.count.jpg
                        cv2.imwrite("training/dataset/User." + str(int(face_id)-1) + '.' + str(count) + ".jpg", gray)
                        cv2image = cv2.cvtColor(img, cv2.COLOR_BGR2RGBA)

                        img_update = ImageTk.PhotoImage(image=Image.fromarray(cv2image))

                        lable_image = Label(newattwindow)
                        lable_image.place(x=1200, y=50)
                        lable_image.place(x=1200, y=50)
                        lable_image.configure(image=img_update)
                        lable_image.image = img_update
                        lable_image.update()
                        # cv2.imshow('Taking Face Samples', img)
                    progressing()
                    k = cv2.waitKey(100) & 0xff
                    if k == 27:
                        break
                    elif count >= 100:
                        break
                cam.release()
                cv2.destroyAllWindows()
                successtakeface()
                training_lable = Label(newattwindow,
                                       text='Training your face',
                                       font=('Montserrat', 12),
                                       fg="#000000",
                                       bg='#f0f0f0')
                training_lable.place(x=400, y=690)
                training = Progressbar(newattwindow, orient=HORIZONTAL, length=1450)
                training.place(x=400, y=720)
                for i in range(0, 1000):
                    training['value'] += 0.01
                    newattwindow.update_idletasks()
                call(["python", "training/face_training.py"])
                for i in range(0, 1000):
                    training['value'] += 0.1
                    newattwindow.update_idletasks()
                successtraining()
                # newattwindow.destroy()

            def successtakeface():
                messagebox.showinfo(title='Done',
                                    message='Completed your face capture!!! Click OK to start the training process')
            def successtraining():
                messagebox.showinfo(title='Done',
                                    message='Completed face training!!!\n Thank you ' + str(name))
                inmain()

        '''---------Hàm điểm danh---------'''
        def attendance():
            wb = load_workbook('../database/namedb.xlsx')
            ws = wb.active
            col = ws['A']
            labels = []
            user=[]
            passw=[]
            '''Database'''
            conn = sqlite3.connect('../database/info.db')
            c = conn.cursor()
            c.execute("SELECT rowid,* FROM info")
            items = c.fetchall()
            for item in items:
                print(item)
                labels.append(item[2])
                user.append(item[3])
                passw.append(item[4])
            print(labels)
            conn.commit()
            conn.close()

            '''---------Hàm điểm danh vào làm---------'''
            def attin():
                def getImagesAndLabels():

                    path = 'training/dataset'
                    imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
                    faceSamples = []
                    ids = []

                    for imagePath in imagePaths:

                        # if there is an error saving any jpegs
                        try:
                            PIL_img = Image.open(imagePath).convert('L')  # convert it to grayscale
                        except:
                            continue
                        img_numpy = np.array(PIL_img, 'uint8')

                        id = int(os.path.split(imagePath)[-1].split(".")[1])
                        faceSamples.append(img_numpy)
                        ids.append(id)
                    return faceSamples, ids

                _, ids = getImagesAndLabels()
                model = training.Model.model((32, 32, 1), len(set(ids)))

                model.load_weights('training/dataset/trained_model.h5')
                model.summary()  # đánh giá độ phù hợp của model

                cascPath = "haarcascade_frontalface_default.xml"
                faceCascade = cv2.CascadeClassifier(cascPath)
                font = cv2.FONT_HERSHEY_SIMPLEX

                def start():
                    cap = cv2.VideoCapture(0)
                    print('here')
                    ret = True
                    clip = []
                    while ret:
                        # read frame by frame
                        ret, frame = cap.read()
                        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
                        img_update = ImageTk.PhotoImage(image=Image.fromarray(cv2image))
                        lable_image = Label(attwindow)
                        lable_image.place(x=770, y=250)
                        lable_image.configure(image=img_update)
                        lable_image.image = img_update
                        lable_image.update()
                        nframe = frame
                        faces = faceCascade.detectMultiScale(
                            frame,
                            scaleFactor=1.1,  # độ scale sau mỗi lần quét
                            minNeighbors=5,
                            # scale và quét ảnh cho đến khi không thể scale được nữa thì lúc này sẽ xuất hiện những khung ảnh trùng nhau, số lần trùng nhau chính là tham số minNeighbors để quyết định cho việc có chọn khung ảnh này là khuôn mặt hay không.
                            minSize=(30, 30))

                        try:
                            (x, y, w, h) = faces[0]
                        except:
                            print("No face")
                            continue
                        frame = frame[y:y + h, x:x + w]
                        frame = cv2.resize(frame, (32, 32))
                        # gray = chuyenanhgray(frame, 'average')
                        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                        # cv2.imshow('result small', frame)
                        c = cv2.waitKey(1)
                        if c & 0xFF == ord('q'):
                            break
                        # gray = gray[np.newaxis,:,:,np.newaxis]
                        gray = gray.reshape(-1, 32, 32, 1).astype('float32') / 255.
                        print(gray.shape)
                        prediction = model.predict(gray)  # dự đoán
                        print("prediction:" + str(prediction))

                        print("\n\n\n\n")
                        print("----------------------------------------------")

                        prediction = prediction.tolist()

                        listv = prediction[0]
                        n = listv.index(max(listv))
                        print("\n")
                        print("----------------------------------------------")
                        print("Highest Probability: " + labels[n] + "==>" + str(prediction[0][n]))
                        # print( "Highest Probability: " + "User " + str(n) + "==>" + str(prediction[0][n]) )
                        global result
                        result = str(labels[n])
                        result_pass=str(passw[n])
                        result_user=str(user[n])
                        if result is None:
                            pass
                        else:
                            print(result)
                            if messagebox.askyesno(title='Attendance Confirmation', message=f'Is you name {result}?'):
                                wb1 = load_workbook('../database/att_db.xlsx')
                                ws1 = wb1.active
                                # name=ws1['A']
                                # timein=ws1['B']
                                now = datetime.now()
                                dtString = now.strftime('%d/%m/%Y,%H:%M:%S')
                                flag = 0
                                for count in range(1, 100):
                                    name = 'A' + str(count)
                                    time_att = 'B' + str(count)
                                    if ws1[name].value is None:
                                        ws1[name] = result
                                        ws1[time_att] = dtString
                                        wb1.save('att_db.xlsx')
                                        wb1.close()
                                        break
                                    else:
                                        pass
                                '''Database'''
                                conn = sqlite3.connect('../database/info.db')
                                c = conn.cursor()
                                c.execute('INSERT INTO Attendance (name,TimeIn,Password,Username) values (?,?,?,?)',(result, dtString,result_pass,result_user))
                                conn.commit()
                                conn.close()
                                success_att()
                                inmain()
                                break
                            else:
                                attendance()
                                break

                        print("----------------------------------------------")
                        print("\n")
                        for (x, y, w, h) in faces:
                            try:
                                cv2.rectangle(nframe, (x, y), (x + w, y + h), (0, 255, 0),
                                              2)  # (0,255,0) hình chữu nhật có màu xanh, độ dày là 2
                                cv2.putText(nframe, str(labels[n]), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
                                # cv2.putText(nframe, str(confidence), (x+5,y+h-5), font, 1, (255,255,0), 1)

                            except:
                                la = 2
                        prediction = np.argmax(model.predict(gray), 1)
                        print(prediction)
                        cv2.imshow('result', nframe)
                        c = cv2.waitKey(1)
                        if c & 0xFF == ord('q'):
                            break

                    cap.release()
                    cv2.destroyAllWindows()

                start()

            '''---------Hàm điểm danh tan làm---------'''
            def attout():
                def getImagesAndLabels():

                    path = 'training/dataset'
                    imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
                    faceSamples = []
                    ids = []

                    for imagePath in imagePaths:

                        # if there is an error saving any jpegs
                        try:
                            PIL_img = Image.open(imagePath).convert('L')  # convert it to grayscale
                        except:
                            continue
                        img_numpy = np.array(PIL_img, 'uint8')

                        id = int(os.path.split(imagePath)[-1].split(".")[1])
                        faceSamples.append(img_numpy)
                        ids.append(id)
                    return faceSamples, ids

                _, ids = getImagesAndLabels()
                model = training.Model.model((32, 32, 1), len(set(ids)))

                model.load_weights('training/dataset/trained_model.h5')
                model.summary()  # đánh giá độ phù hợp của model

                cascPath = "haarcascade_frontalface_default.xml"
                faceCascade = cv2.CascadeClassifier(cascPath)
                font = cv2.FONT_HERSHEY_SIMPLEX

                def start():
                    cap = cv2.VideoCapture(0)
                    print('here')
                    ret = True

                    clip = []
                    while ret:
                        # read frame by frame
                        ret, frame = cap.read()
                        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
                        img_update = ImageTk.PhotoImage(image=Image.fromarray(cv2image))
                        lable_image = Label(attwindow)
                        lable_image.place(x=770, y=250)
                        lable_image.configure(image=img_update)
                        lable_image.image = img_update
                        lable_image.update()
                        nframe = frame
                        faces = faceCascade.detectMultiScale(
                            frame,
                            scaleFactor=1.1,  # độ scale sau mỗi lần quét
                            minNeighbors=5,
                            # scale và quét ảnh cho đến khi không thể scale được nữa thì lúc này sẽ xuất hiện những khung ảnh trùng nhau, số lần trùng nhau chính là tham số minNeighbors để quyết định cho việc có chọn khung ảnh này là khuôn mặt hay không.
                            minSize=(30, 30))

                        try:
                            (x, y, w, h) = faces[0]
                        except:
                            print("No face")
                            continue
                        frame = frame[y:y + h, x:x + w]
                        frame = cv2.resize(frame, (32, 32))
                        # gray = chuyenanhgray(frame, 'lightness')
                        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                        cv2.imshow('result small', frame)
                        c = cv2.waitKey(1)
                        if c & 0xFF == ord('q'):
                            break

                        # gray = gray[np.newaxis,:,:,np.newaxis]
                        gray = gray.reshape(-1, 32, 32, 1).astype('float32') / 255.
                        print(gray.shape)
                        prediction = model.predict(gray)  # dự đoán
                        print("prediction:" + str(prediction))

                        print("\n\n\n\n")
                        print("----------------------------------------------")

                        prediction = prediction.tolist()

                        listv = prediction[0]
                        n = listv.index(max(listv))
                        print("\n")
                        print("----------------------------------------------")
                        print("Highest Probability: " + labels[n] + "==>" + str(prediction[0][n]))
                        # print( "Highest Probability: " + "User " + str(n) + "==>" + str(prediction[0][n]) )
                        global result
                        result = str(labels[n])
                        result_pass = str(passw[n])
                        result_user = str(user[n])
                        if result is None:
                            pass
                        else:
                            print(result)
                            if messagebox.askyesno(title='Attendance Confirmation', message=f'Is you name {result}?'):
                                wb1 = load_workbook('../database/att_db.xlsx')
                                ws1 = wb1.active
                                # name=ws1['A']
                                # timein=ws1['B']
                                now = datetime.now()
                                dtString = now.strftime('%d/%m/%Y,%H:%M:%S')
                                flag = 0
                                for count in range(1, 100):
                                    name = 'A' + str(count)
                                    time_att = 'C' + str(count)
                                    if ws1[name].value is None:
                                        ws1[name] = result
                                        ws1[time_att] = dtString
                                        wb1.save('../database/att_db.xlsx')
                                        wb1.close()
                                        break
                                    else:
                                        pass
                                '''Database'''
                                conn = sqlite3.connect('../database/info.db')
                                c = conn.cursor()
                                c.execute('INSERT INTO Attendance (name,TimeOut,Password,Username) values (?,?,?,?)',(result, dtString,result_pass,result_user))
                                conn.commit()
                                conn.close()
                                success_att()
                                inmain()
                                break
                            else:
                                attendance()
                                break

                        print("----------------------------------------------")
                        print("\n")
                        for (x, y, w, h) in faces:
                            try:
                                cv2.rectangle(nframe, (x, y), (x + w, y + h), (0, 255, 0),
                                              2)  # (0,255,0) hình chữu nhật có màu xanh, độ dày là 2
                                cv2.putText(nframe, str(labels[n]), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
                                # cv2.putText(nframe, str(confidence), (x+5,y+h-5), font, 1, (255,255,0), 1)

                            except:
                                la = 2
                        prediction = np.argmax(model.predict(gray), 1)
                        print(prediction)
                        # cv2.imshow('result', nframe)
                        c = cv2.waitKey(1)
                        if c & 0xFF == ord('q'):
                            break

                    cap.release()
                    cv2.destroyAllWindows()


                start()

            def success_att():
                messagebox.showinfo(title='Done',
                                    message='Successful Attendance!!! Thanks ' + result)

            attwindow = Canvas(window, width=width_value, height=height_value)
            attwindow.place(x=0, y=0)
            '''Menu'''

            menu_side = Label(attwindow, bg='#1C2A3A', padx=150, pady=1080)
            menu_side.place(x=0, y=0)
            logo = ImageTk.PhotoImage(Image.open('../pic/logo77.jpg'))
            lg = Label(attwindow, image=logo)
            lg.image = logo
            lg.place(x=27, y=27)
            Label(attwindow, text='Face Attendance', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=20)
            Label(attwindow, text='System', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=50)
            Label(attwindow, text='Group 7', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=80)

            home = Button(
                window,
                text='Home',
                command=inmain,
                font=('Montserrat', 15),
                height='2',
                width='22',
                fg='#ffffff',
                bg='#1C2A3A',
                bd=0
            )
            newatt = Button(window,
                            text='Register New Employee',
                            command=newattendace,
                            font=('Montserrat', 15),
                            height='2',
                            width='22',
                            fg='#ffffff',
                            bg='#1C2A3A',
                            bd=0
                            )
            att = Button(window,
                         text='Mark Attendance',
                         font=('Montserrat', 15),
                         command=attendance,
                         height='2',
                         width='23',
                         fg='#ffffff',
                         bg='#1FB29E',
                         bd=0
                         )
            result_att = Button(window,
                            text='Result Attendance',
                            font=('Montserrat', 15),
                            command=result_attendance,
                            height='2',
                            width='22',
                            fg='#ffffff',
                            bg='#1C2A3A',
                            bd=0
                            )
            home.place(x=0, y=150)
            newatt.place(x=0, y=250)
            att.place(x=0, y=350)
            result_att.place(x=0, y=450)


            '''Realtime display'''
            def my_time():
                import time
                result_time = time.localtime()
                wday=['MONDAY','TUESDAY','WEDNESDAY','THURSDAY','FRIDAY','SATURDAY','SUNDAY']
                month=['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']
                for i in range(0,7):
                    if result_time.tm_wday == i:
                        dayofweek=wday[i]
                for j in range(0,12):
                    if result_time.tm_mon == j:
                        month_label=month[j-1]
                time=str(format(result_time.tm_hour,"0>2,d"))+':'+str(format(result_time.tm_min,"0>2,d"))+':'+str(format(result_time.tm_sec, "0>2,d"))
                day=str(month_label)+' '+str(result_time.tm_mday)+', '+str(result_time.tm_year)
                Label(attwindow,text=dayofweek,font=('Montserrat', 30, 'bold'),fg="#000000",bg='#f0f0f0').place(x=1000,y=0)
                Label(attwindow, text=time, font=('Montserrat', 50, 'bold'), fg="#000000", bg='#f0f0f0',height=1,width=10).place(x=860,y=50)
                Label(attwindow, text=day, font=('Montserrat Regular', 20), fg="#000000", bg='#f0f0f0').place(x=1000,y=145)
                attwindow.after(1000,my_time)
            my_time()
            '''IN'''
            att_in = Button(attwindow,
                            text='Mark Your Attendance - IN',
                            font=('Montserrat', 15),
                            command=attin,
                            height='2',
                            width='25',
                            fg='#9aa9a9',
                            bg='#252527',
                            )
            '''OUT'''
            att_in.place(x=400, y=850)
            att_out = Button(attwindow,
                             text='Mark Your Attendance - OUT',
                             font=('Montserrat', 15),
                             command=attout,
                             height='2',
                             width='25',
                             fg='#9aa9a9',
                             bg='#252527',
                             )
            att_out.place(x=1470, y=850)
            '''Picture'''
            cam_photo = Image.open('../pic/Attendance_Picture.png')
            cam_photo = ImageTk.PhotoImage(cam_photo)
            lable_image = Label(attwindow, image=cam_photo)
            lable_image.image = cam_photo
            lable_image.place(x=770, y=250)

        '''---------Hàm tra cứu kết quả---------'''
        def result_attendance():
            result_window = Canvas(window, width=width_value, height=height_value)
            result_window.place(x=0, y=0)
            menu_side = Label(result_window, bg='#1C2A3A', padx=150, pady=1080)
            menu_side.place(x=0, y=0)
            logo = ImageTk.PhotoImage(Image.open('../pic/logo77.jpg'))
            lg = Label(result_window, image=logo)
            lg.image = logo
            lg.place(x=27, y=27)
            Label(result_window, text='Face Attendance', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=20)
            Label(result_window, text='System', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=50)
            Label(result_window, text='Group 7', font=('Montserrat', 15), bg='#1C2A3A', fg='#ffffff').place(x=110, y=80)
            home = Button(
                window,
                text='Home',
                command=inmain,
                font=('Montserrat', 15),
                height='2',
                width='22',
                fg='#ffffff',
                bg='#1C2A3A',
                bd=0
            )
            newatt = Button(window,
                            text='Register New Employee',
                            command=newattendace,
                            font=('Montserrat', 15),
                            height='2',
                            width='22',
                            fg='#ffffff',
                            bg='#1C2A3A',
                            bd=0
                            )
            att = Button(window,
                         text='Mark Attendance',
                         font=('Montserrat', 15),
                         command=attendance,
                         height='2',
                         width='22',
                         fg='#ffffff',
                         bg='#1C2A3A',
                         bd=0
                         )
            result_att = Button(window,
                                text='Result Attendance',
                                font=('Montserrat', 15),
                                command=result_attendance,
                                height='2',
                                width='23',
                                fg='#ffffff',
                                bg='#1FB29E',
                                bd=0
                                )
            home.place(x=0, y=150)
            newatt.place(x=0, y=250)
            att.place(x=0, y=350)
            result_att.place(x=0, y=450)

            '''--------Hàm đăng nhập---------'''
            def login():
                '''------------------Bang du lieu---------------'''
                conn = sqlite3.connect('../database/info.db')
                '''Tao cursor'''
                c = conn.cursor()
                c.execute("SELECT rowid,* FROM info")
                items = c.fetchall()
                flag_username=0
                flag_password=0
                check_flag=0
                print(str(username_entry.get()))
                print(str(password_entry.get()))
                Label(result_window, text='*Wrong Username', font=('Montserrat Regular', 10), fg="#f0f0f0").place(x=665,y=65)
                Label(result_window, text='*Wrong Password', font=('Montserrat Regular', 10), fg="#f0f0f0").place(x=670,y=185)
                Label(result_window, bg='#f0f0f0', fg='#000000', height=100, width=500, borderwidth=0).place(x=420, y=350)

                for item in items:
                    print(item[3])
                    print(item[4])
                    if str(item[3]) == str(username_entry.get()):
                        flag_username=1
                        if str(item[4]) == str(password_entry.get()):
                            flag_password=1
                            info_c=conn.cursor()
                            info_c.execute("SELECT rowid,* FROM Attendance")
                            items_attendance=info_c.fetchall()
                            data_counter=2
                            y_result = 350
                            Label(result_window, text='Attendance Record', font=('Montserrat', 15), bg='#ffffff', fg='#000000',
                                  height=2, width=100, anchor='w', borderwidth=1, relief="ridge").place(x=420, y=y_result)
                            Label(result_window, text='Name', font=('Montserrat', 15), bg='#f0f0f0', fg='#000000', height=2,
                                  width=30, relief="ridge").place(x=420, y=y_result + 53)
                            Label(result_window, text='Time In', font=('Montserrat', 15), bg='#f0f0f0', fg='#000000', height=2,
                                  width=35, relief="ridge").place(x=810, y=y_result + 53)
                            Label(result_window, text='Time Out', font=('Montserrat', 15), bg='#f0f0f0', fg='#000000', height=2,
                                  width=35, relief="ridge").place(x=1263, y=y_result + 53)
                            for item_att in reversed(items_attendance):
                                if item_att[4]==item[3] and item_att[5]==item[4]:
                                    print(item_att)
                                    Label(result_window, text=item_att[1], font=('Montserrat Regular', 15), bg='#ffffff', fg='#000000',
                                          height=2, width=30,
                                          relief="ridge").place(x=420, y=y_result + 53 * data_counter)

                                    Label(result_window, text=item_att[2], font=('Montserrat Regular', 15), bg='#ffffff', fg='#000000',
                                          height=2,
                                          width=35, relief="ridge").place(x=810, y=y_result + 53 * data_counter)

                                    Label(result_window, text=item_att[3], font=('Montserrat Regular', 15), bg='#ffffff', fg='#000000',
                                          height=2,
                                          width=35, relief="ridge").place(x=1263, y=y_result + 53 * data_counter)
                                    data_counter+=1
                                    if data_counter == 10:
                                        break

                if flag_username==0:
                    Label(result_window, text='*Wrong Username', font=('Montserrat Regular', 10), fg="#ff0000").place(x=665, y=65)
                if flag_password==0:
                    Label(result_window, text='*Wrong Password', font=('Montserrat Regular', 10), fg="#ff0000").place(x=670, y=185)

                conn.commit()
                conn.close()
            '''---------------------Login-------------------'''
            username_lable = Label(result_window,
                             text='Username:',
                             font=('Montserrat Light', 20),
                             fg="#000000",
                             bg='#f0f0f0')
            password_lable = Label(result_window,
                               text='Password:',
                               font=('Montserrat Light', 20),
                               fg="#000000",
                               bg='#f0f0f0')
            username_lable.place(x=400, y=50)
            password_lable.place(x=400, y=170)

            username_entry = Entry(result_window,
                             font=('Montserrat', 20),
                             bg='#ffffff',
                             fg='#000000')
            password_entry = Entry(result_window,
                               font=('Montserrat', 20),
                               bg='#ffffff',
                               fg='#000000',show='*')
            username_entry.place(x=405, y=95)
            password_entry.place(x=405, y=215)

            '''---Button---'''

            login_in = Button(result_window,
                                text='Login',
                                command=login,
                                font=('Montserrat', 15),
                                height='1',
                                width='15',
                                fg='#ffffff',
                                bg='#2B3647',
                                )
            login_in.place(x=585, y=280)




        '''Home'''
        home=Button(window,text='Home',command=inmain,font=('Montserrat', 15),height='2',width='23',fg='#ffffff',bg='#1FB29E',bd=0)
        newatt=Button(window,text='Register New Employee',command=newattendace,font=('Montserrat',15),height='2',width='22',fg='#ffffff',bg='#1C2A3A',bd=0)
        att=Button(window,text='Mark Attendance',font=('Montserrat',15),command=attendance,height='2',width='22',fg='#ffffff',bg='#1C2A3A',bd=0)
        result_att=Button(window,text='Result Attendance',font=('Montserrat',15),command=result_attendance,height='2',width='22',fg='#ffffff',bg='#1C2A3A',bd=0)
        home.place(x=0,y=150)
        newatt.place(x=0,y=250)
        att.place(x=0,y=350)
        result_att.place(x=0,y=450)
        #lable.place(x=0,y=0)

    inmain()
    window.mainloop()
mainfunc()
