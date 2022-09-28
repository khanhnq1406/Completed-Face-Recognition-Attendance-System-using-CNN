from tkinter import *
from tkinter import messagebox
import numpy as np
import cv2
import os
import h5py
import dlib
from imutils import face_utils
from keras.models import load_model
import sys
from keras.models import Sequential
from keras.layers import Conv2D, MaxPooling2D,Dropout
from keras.layers import Dense, Activation, Flatten
from PIL import Image
from datetime import datetime
import training.Model
from training.Model import model
from openpyxl import Workbook,load_workbook
wb=load_workbook('../database/namedb.xlsx')
ws=wb.active
col=ws['A']
labels=[]
for cell in col:
    labels.append(cell.value)

def attendance():
    attwindow = Tk()
    attwindow.title("Mark Attendance")
    attwindow.geometry("960x540")
    attwindow.config(background="#323436")
    lable = Label(attwindow,
                  text='Mark Attendance',
                  font=('Montserrat', 30, 'bold'),
                  fg="#9aa9a9",
                  bg='#323436')
    lable.pack()
    def attin():
        def getImagesAndLabels():

            path = 'training/dataset'
            imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
            faceSamples = []
            ids = []

            for imagePath in imagePaths:

                # if there is an error saving any jpegs
                try:
                    PIL_img = Image.open(imagePath).convert('L')  # convert it to grayscale
                except:
                    continue
                img_numpy = np.array(PIL_img, 'uint8')

                id = int(os.path.split(imagePath)[-1].split(".")[1])
                faceSamples.append(img_numpy)
                ids.append(id)
            return faceSamples, ids

        _, ids = getImagesAndLabels()
        model = training.Model.model((32, 32, 1), len(set(ids)))

        model.load_weights ('training/dataset/trained_model.h5')
        model.summary()  # đánh giá độ phù hợp của model

        cascPath = "haarcascade_frontalface_default.xml"
        faceCascade = cv2.CascadeClassifier(cascPath)
        font = cv2.FONT_HERSHEY_SIMPLEX

        def start():
            cap = cv2.VideoCapture(0)
            print('here')
            ret = True

            clip = []
            while ret:
                # read frame by frame
                ret, frame = cap.read()
                nframe = frame
                faces = faceCascade.detectMultiScale(
                    frame,
                    scaleFactor=1.1,  # độ scale sau mỗi lần quét
                    minNeighbors=5,
                    # scale và quét ảnh cho đến khi không thể scale được nữa thì lúc này sẽ xuất hiện những khung ảnh trùng nhau, số lần trùng nhau chính là tham số minNeighbors để quyết định cho việc có chọn khung ảnh này là khuôn mặt hay không.
                    minSize=(30, 30))

                try:
                    (x, y, w, h) = faces[0]
                except:
                    print("No face")
                    continue
                frame = frame[y:y + h, x:x + w]
                frame = cv2.resize(frame, (32, 32))
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                cv2.imshow('result small', frame)
                c = cv2.waitKey(1)
                if c & 0xFF == ord('q'):
                    break

                # gray = gray[np.newaxis,:,:,np.newaxis]
                gray = gray.reshape(-1, 32, 32, 1).astype('float32') / 255.
                print(gray.shape)
                prediction = model.predict(gray)  # dự đoán
                print("prediction:" + str(prediction))

                print("\n\n\n\n")
                print("----------------------------------------------")

                prediction = prediction.tolist()

                listv = prediction[0]
                n = listv.index(max(listv))
                print("\n")
                print("----------------------------------------------")
                print("Highest Probability: " + labels[n] + "==>" + str(prediction[0][n]))
                # print( "Highest Probability: " + "User " + str(n) + "==>" + str(prediction[0][n]) )
                global result
                result = str(labels[n])
                if result is None:
                    pass
                else:
                    print(result)
                    if messagebox.askyesno(title='Attendance Confirmation',message=f'Is you name {result}?'):
                        wb1=load_workbook('../database/att_db.xlsx')
                        ws1=wb1.active
                        # name=ws1['A']
                        # timein=ws1['B']
                        now=datetime.now()
                        dtString=now.strftime('%d/%m/%Y,%H:%M:%S')
                        flag=0
                        for count in range(1, 100):
                            name = 'A' + str(count)
                            time_att='B' +str(count)
                            if ws1[name].value is None:
                                ws1[name] = result
                                ws1[time_att]=dtString
                                wb1.save('../database/att_db.xlsx')
                                wb1.close()
                                break
                            else:
                                pass


                        break
                    else:
                        start()


                print("----------------------------------------------")
                print("\n")
                for (x, y, w, h) in faces:
                    try:
                        cv2.rectangle(nframe, (x, y), (x + w, y + h), (0, 255, 0),
                                      2)  # (0,255,0) hình chữu nhật có màu xanh, độ dày là 2
                        cv2.putText(nframe, str(labels[n]), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
                        # cv2.putText(nframe, str(confidence), (x+5,y+h-5), font, 1, (255,255,0), 1)

                    except:
                        la = 2
                prediction = np.argmax(model.predict(gray), 1)
                print(prediction)
                cv2.imshow('result', nframe)
                c = cv2.waitKey(1)
                if c & 0xFF == ord('q'):
                    break

            cap.release()
            cv2.destroyAllWindows()

        start()
    def attout():
        def getImagesAndLabels():

            path = 'training/dataset'
            imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
            faceSamples = []
            ids = []

            for imagePath in imagePaths:

                # if there is an error saving any jpegs
                try:
                    PIL_img = Image.open(imagePath).convert('L')  # convert it to grayscale
                except:
                    continue
                img_numpy = np.array(PIL_img, 'uint8')

                id = int(os.path.split(imagePath)[-1].split(".")[1])
                faceSamples.append(img_numpy)
                ids.append(id)
            return faceSamples, ids

        _, ids = getImagesAndLabels()
        model = training.Model.model((32, 32, 1), len(set(ids)))

        model.load_weights ('training/dataset/trained_model.h5')
        model.summary()  # đánh giá độ phù hợp của model

        cascPath = "haarcascade_frontalface_default.xml"
        faceCascade = cv2.CascadeClassifier(cascPath)
        font = cv2.FONT_HERSHEY_SIMPLEX

        def start():
            cap = cv2.VideoCapture(0)
            print('here')
            ret = True

            clip = []
            while ret:
                # read frame by frame
                ret, frame = cap.read()
                nframe = frame
                faces = faceCascade.detectMultiScale(
                    frame,
                    scaleFactor=1.1,  # độ scale sau mỗi lần quét
                    minNeighbors=5,
                    # scale và quét ảnh cho đến khi không thể scale được nữa thì lúc này sẽ xuất hiện những khung ảnh trùng nhau, số lần trùng nhau chính là tham số minNeighbors để quyết định cho việc có chọn khung ảnh này là khuôn mặt hay không.
                    minSize=(30, 30))

                try:
                    (x, y, w, h) = faces[0]
                except:
                    print("No face")
                    continue
                frame = frame[y:y + h, x:x + w]
                frame = cv2.resize(frame, (32, 32))
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                cv2.imshow('result small', frame)
                c = cv2.waitKey(1)
                if c & 0xFF == ord('q'):
                    break

                # gray = gray[np.newaxis,:,:,np.newaxis]
                gray = gray.reshape(-1, 32, 32, 1).astype('float32') / 255.
                print(gray.shape)
                prediction = model.predict(gray)  # dự đoán
                print("prediction:" + str(prediction))

                print("\n\n\n\n")
                print("----------------------------------------------")

                prediction = prediction.tolist()

                listv = prediction[0]
                n = listv.index(max(listv))
                print("\n")
                print("----------------------------------------------")
                print("Highest Probability: " + labels[n] + "==>" + str(prediction[0][n]))
                # print( "Highest Probability: " + "User " + str(n) + "==>" + str(prediction[0][n]) )
                global result
                result = str(labels[n])
                if result is None:
                    pass
                else:
                    print(result)
                    if messagebox.askyesno(title='Attendance Confirmation',message=f'Is you name {result}?'):
                        wb1=load_workbook('../database/att_db.xlsx')
                        ws1=wb1.active
                        # name=ws1['A']
                        # timein=ws1['B']
                        now=datetime.now()
                        dtString=now.strftime('%d/%m/%Y,%H:%M:%S')
                        flag=0
                        for count in range(1, 100):
                            name = 'A' + str(count)
                            time_att='C' +str(count)
                            if ws1[name].value is None:
                                ws1[name] = result
                                ws1[time_att]=dtString
                                wb1.save('../database/att_db.xlsx')
                                wb1.close()
                                break
                            else:
                                pass


                        break
                    else:
                        start()


                print("----------------------------------------------")
                print("\n")
                for (x, y, w, h) in faces:
                    try:
                        cv2.rectangle(nframe, (x, y), (x + w, y + h), (0, 255, 0),
                                      2)  # (0,255,0) hình chữu nhật có màu xanh, độ dày là 2
                        cv2.putText(nframe, str(labels[n]), (x + 5, y - 5), font, 1, (255, 255, 255), 2)
                        # cv2.putText(nframe, str(confidence), (x+5,y+h-5), font, 1, (255,255,0), 1)

                    except:
                        la = 2
                prediction = np.argmax(model.predict(gray), 1)
                print(prediction)
                cv2.imshow('result', nframe)
                c = cv2.waitKey(1)
                if c & 0xFF == ord('q'):
                    break

            cap.release()
            cv2.destroyAllWindows()

        start()
    att_in = Button(attwindow,
                    text='Mark Your Attendance - IN',
                    font=('Montserrat', 15),
                    command=attin,
                    height='3',
                    width='25',
                    fg='#9aa9a9',
                    bg='#252527',
                    )
    att_in.place(x=100,y=200)
    att_out = Button(attwindow,
                    text='Mark Your Attendance - OUT',
                    font=('Montserrat', 15),
                    command=attout,
                    height='3',
                    width='25',
                    fg='#9aa9a9',
                    bg='#252527',
                    )
    att_out.place(x=530, y=200)

